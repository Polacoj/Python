"""
validar_operadores.py
=====================
Objetivo:
  1. Leer el diccionario maestro (PRODUCTIVIDAD_GENERAL.xlsx - hoja PERSONAL)
  2. Leer los operadores de ADIA (dia02.xlsx - Hoja1, col OPERADOR)
  3. Leer los operadores de ACD  (CONTROL DIARIO MAYO 2026.xlsx - Mayo 2026,
     cols OPERADOR / COLABORADORES 1-3 / SAE / SISEP)
  4. Para cada nombre encontrado intentar una coincidencia fuzzy contra el dict.
  5. Mostrar los NO-MATCH con sugerencias y permitir resolución con GUI.
  6. Actualizar el Excel del diccionario maestro con las altas aprobadas.

Librerías:
  - pandas      → leer/escribir Excel
  - rapidfuzz   → comparación fuzzy de cadenas (moderna, muy rápida)
  - unicodedata → normalización de acentos
  - openpyxl    → escribir de vuelta en el xlsx preservando formato
  - PyQt6       → interfaz gráfica para la revisión interactiva
"""

import os
import re
import unicodedata
import pandas as pd
from rapidfuzz import process, fuzz
from openpyxl import load_workbook
from PyQt6.QtWidgets import (
    QApplication,
    QDialog,
    QFileDialog,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QListWidget,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

DICT_FILE = "PRODUCTIVIDAD_GENERAL.xlsx"
DICT_SHEET = "PERSONAL"
DICT_COL_ID = "LP / DNI"
DICT_COL_NM = "Nombre Completo"

ADIA_FILE = "dia02.xlsx"
ADIA_SHEET = "Hoja1"
ADIA_COLS = ["OPERADOR"]

ACD_FILE = "CONTROL DIARIO MAYO 2026.xlsx"
ACD_SHEET = "Mayo 2026"
ACD_HEADER_ROW = 1
ACD_COLS = ["OPERADOR", "COLABORADORES 1", "COLABORADORES 2",
            "COLABORADORES 3", "SAE.1", "SISEP"]

UMBRAL_MATCH = 82


def normalizar(texto: str) -> str:
    if not isinstance(texto, str):
        return ""
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    texto = texto.upper().strip()
    texto = re.sub(r"\s+", " ", texto)
    return texto


def tokens_ordenados(texto: str) -> str:
    palabras = normalizar(texto).split()
    return " ".join(sorted(palabras))


def cargar_diccionario(filepath: str, sheet: str) -> pd.DataFrame:
    df = pd.read_excel(filepath, sheet_name=sheet, dtype={DICT_COL_ID: str})
    df = df[[DICT_COL_ID, DICT_COL_NM]].dropna(subset=[DICT_COL_NM])
    df["nombre_norm"] = df[DICT_COL_NM].apply(tokens_ordenados)
    return df.reset_index(drop=True)


def extraer_nombres_adia(filepath: str, sheet: str, cols: list) -> set:
    df = pd.read_excel(filepath, sheet_name=sheet)
    nombres = set()
    for col in cols:
        if col in df.columns:
            nombres.update(df[col].dropna().astype(str).str.strip().unique())
    return nombres - {""}


def extraer_nombres_acd(filepath: str, sheet: str, header_row: int,
                        cols: list) -> set:
    df = pd.read_excel(filepath, sheet_name=sheet, header=header_row)
    nombres = set()
    for col in cols:
        if col in df.columns:
            nombres.update(df[col].dropna().astype(str).str.strip().unique())
    return nombres - {""}


def buscar_en_diccionario(nombre_raw: str, diccionario: pd.DataFrame,
                          umbral: int = UMBRAL_MATCH):
    clave = tokens_ordenados(nombre_raw)
    if not clave:
        return None, 0

    resultado = process.extractOne(
        clave,
        diccionario["nombre_norm"],
        scorer=fuzz.token_sort_ratio,
        score_cutoff=0,
    )
    if resultado is None:
        return None, 0

    _, score, idx = resultado
    if score >= umbral:
        return diccionario.iloc[idx], score
    return None, score


def guardar_nuevas_entradas(filepath: str, sheet: str, nuevas: list,
                            diccionario: pd.DataFrame):
    if not nuevas:
        return

    wb = load_workbook(filepath)
    ws = wb[sheet]
    ultima_fila = ws.max_row

    headers = {ws.cell(row=1, column=c).value: c
               for c in range(1, ws.max_column + 1)}
    col_id = headers.get(DICT_COL_ID)
    col_nm = headers.get(DICT_COL_NM)
    if col_id is None or col_nm is None:
        raise ValueError(
            f"No se encontraron las columnas '{DICT_COL_ID}' y/o '{DICT_COL_NM}' en la hoja."
        )

    for entrada in nuevas:
        ultima_fila += 1
        ws.cell(row=ultima_fila, column=col_id, value=entrada[DICT_COL_ID])
        ws.cell(row=ultima_fila, column=col_nm, value=entrada[DICT_COL_NM])

    wb.save(filepath)


class MismatchDialog(QDialog):
    def __init__(self, nombre_raw: str, sugerencias: list, parent=None):
        super().__init__(parent)
        self.nombre_raw = nombre_raw
        self.sugerencias = sugerencias
        self.selected_action = None
        self.selected_suggestion = None
        self.new_entry = None
        self.setWindowTitle("Resolver no coincidencia")
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(f"<b>Nombre encontrado:</b> {self.nombre_raw}"))
        layout.addWidget(QLabel("Sugerencias en el diccionario:"))

        self.suggestions_list = QListWidget()
        if self.sugerencias:
            for fila, score in self.sugerencias:
                self.suggestions_list.addItem(
                    f"[{fila[DICT_COL_ID]}] {fila[DICT_COL_NM]} ({score:.0f}%)"
                )
        else:
            self.suggestions_list.addItem("Sin sugerencias cercanas")
            self.suggestions_list.setEnabled(False)
        layout.addWidget(self.suggestions_list)

        group = QGroupBox("Agregar nuevo al diccionario")
        form = QFormLayout(group)
        self.id_input = QLineEdit()
        self.name_input = QLineEdit()
        form.addRow("LP / DNI:", self.id_input)
        form.addRow("Nombre canónico:", self.name_input)
        layout.addWidget(group)

        botones = QHBoxLayout()
        assign_btn = QPushButton("Asignar sugerencia")
        add_btn = QPushButton("Agregar nuevo")
        skip_btn = QPushButton("Saltar")
        botones.addWidget(assign_btn)
        botones.addWidget(add_btn)
        botones.addWidget(skip_btn)
        layout.addLayout(botones)

        assign_btn.clicked.connect(self.assign_suggestion)
        add_btn.clicked.connect(self.add_new_entry)
        skip_btn.clicked.connect(self.skip)

    def assign_suggestion(self):
        if not self.sugerencias:
            QMessageBox.warning(self, "Selección requerida",
                                "No hay sugerencias disponibles para asignar.")
            return
        idx = self.suggestions_list.currentRow()
        if idx < 0 or idx >= len(self.sugerencias):
            QMessageBox.warning(self, "Selección requerida",
                                "Seleccioná un candidato válido de la lista.")
            return
        self.selected_action = "assign"
        self.selected_suggestion = self.sugerencias[idx][0]
        self.accept()

    def add_new_entry(self):
        nuevo_id = self.id_input.text().strip()
        nuevo_nm = self.name_input.text().strip()
        if not nuevo_id or not nuevo_nm:
            QMessageBox.warning(self, "Datos incompletos",
                                "Ingresá LP/DNI y Nombre canónico para agregar.")
            return
        self.selected_action = "add"
        self.new_entry = {DICT_COL_ID: nuevo_id, DICT_COL_NM: nuevo_nm}
        self.accept()

    def skip(self):
        self.selected_action = "skip"
        self.accept()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Validador de Operadores")
        self.resize(900, 640)
        self.setup_ui()

    def setup_ui(self):
        central = QWidget()
        layout = QVBoxLayout(central)

        def file_row(label_text, line_edit, button):
            row = QHBoxLayout()
            row.addWidget(QLabel(label_text))
            row.addWidget(line_edit)
            row.addWidget(button)
            return row

        self.dict_path = QLineEdit(DICT_FILE)
        self.adia_path = QLineEdit(ADIA_FILE)
        self.acd_path = QLineEdit(ACD_FILE)

        dict_btn = QPushButton("Seleccionar")
        adia_btn = QPushButton("Seleccionar")
        acd_btn = QPushButton("Seleccionar")
        dict_btn.clicked.connect(lambda: self.select_file(self.dict_path))
        adia_btn.clicked.connect(lambda: self.select_file(self.adia_path))
        acd_btn.clicked.connect(lambda: self.select_file(self.acd_path))

        layout.addLayout(file_row("Diccionario maestro:", self.dict_path, dict_btn))
        layout.addLayout(file_row("Archivo ADIA:", self.adia_path, adia_btn))
        layout.addLayout(file_row("Archivo ACD:", self.acd_path, acd_btn))

        self.start_button = QPushButton("Iniciar validación")
        self.start_button.clicked.connect(self.on_start)
        layout.addWidget(self.start_button)

        layout.addWidget(QLabel("Registro de ejecución:"))
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        layout.addWidget(self.log_text)

        self.setCentralWidget(central)

    def select_file(self, line_edit: QLineEdit):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar archivo Excel",
            os.getcwd(),
            "Archivos Excel (*.xlsx *.xlsm);;Todos los archivos (*)",
        )
        if path:
            line_edit.setText(path)

    def append_log(self, mensaje: str):
        self.log_text.append(mensaje)
        QApplication.processEvents()

    def on_start(self):
        self.start_button.setEnabled(False)
        self.log_text.clear()
        try:
            self.run_validation()
        except Exception as exc:
            self.append_log(f"ERROR: {exc}")
            QMessageBox.critical(self, "Error", str(exc))
        finally:
            self.start_button.setEnabled(True)

    def run_validation(self):
        dict_path = self.dict_path.text().strip() or DICT_FILE
        adia_path = self.adia_path.text().strip() or ADIA_FILE
        acd_path = self.acd_path.text().strip() or ACD_FILE

        self.append_log("Iniciando validación de operadores...")
        self.append_log(f"  Diccionario: {dict_path}")
        self.append_log(f"  ADIA: {adia_path}")
        self.append_log(f"  ACD: {acd_path}")

        if not os.path.exists(dict_path):
            raise FileNotFoundError(f"No se encontró el diccionario: {dict_path}")
        if not os.path.exists(adia_path):
            raise FileNotFoundError(f"No se encontró ADIA: {adia_path}")
        if not os.path.exists(acd_path):
            raise FileNotFoundError(f"No se encontró ACD: {acd_path}")

        diccionario = cargar_diccionario(dict_path, DICT_SHEET)
        self.append_log(f"Carga completada: {len(diccionario)} registros en el diccionario.")

        nombres_adia = extraer_nombres_adia(adia_path, ADIA_SHEET, ADIA_COLS)
        self.append_log(f"ADIA: {len(nombres_adia)} nombres únicos extraídos.")

        nombres_acd = extraer_nombres_acd(acd_path, ACD_SHEET,
                                          ACD_HEADER_ROW, ACD_COLS)
        self.append_log(f"ACD: {len(nombres_acd)} nombres únicos extraídos.")

        no_match_adia = self.clasificar(nombres_adia, diccionario, "ADIA")
        no_match_acd = self.clasificar(nombres_acd, diccionario, "ACD")

        todas_nuevas = []

        if no_match_adia:
            nuevas_adia = self.resolver_no_matches(no_match_adia, diccionario, "ADIA")
            todas_nuevas.extend(nuevas_adia)
            for entrada in nuevas_adia:
                nueva_fila = pd.DataFrame([{
                    DICT_COL_ID: entrada[DICT_COL_ID],
                    DICT_COL_NM: entrada[DICT_COL_NM],
                    "nombre_norm": tokens_ordenados(entrada[DICT_COL_NM]),
                }])
                diccionario = pd.concat([diccionario, nueva_fila], ignore_index=True)

        if no_match_acd:
            nuevas_acd = self.resolver_no_matches(no_match_acd, diccionario, "ACD")
            todas_nuevas.extend(nuevas_acd)

        if todas_nuevas:
            guardar_nuevas_entradas(dict_path, DICT_SHEET, todas_nuevas, diccionario)
            self.append_log(f"Se guardaron {len(todas_nuevas)} nueva(s) entrada(s) en el diccionario.")
        else:
            self.append_log("No se agregaron nuevas entradas.")

        self.append_log("Validación finalizada.")
        QMessageBox.information(
            self,
            "Finalizado",
            f"Validación finalizada. Nuevas entradas guardadas: {len(todas_nuevas)}",
        )

    def clasificar(self, nombres_set: set, diccionario: pd.DataFrame, fuente: str) -> list:
        no_match = []
        match_ok = []

        for nombre in sorted(nombres_set):
            fila_match, score = buscar_en_diccionario(nombre, diccionario)
            if fila_match is not None:
                match_ok.append((nombre, fila_match[DICT_COL_NM], score))
            else:
                top3_raw = process.extract(
                    tokens_ordenados(nombre),
                    diccionario["nombre_norm"],
                    scorer=fuzz.token_sort_ratio,
                    limit=3,
                )
                sugerencias = [
                    (diccionario.iloc[idx], s)
                    for (_, s, idx) in top3_raw if s >= 50
                ]
                no_match.append((nombre, sugerencias))

        self.append_log(f"{fuente}: {len(match_ok)} coincidencias, {len(no_match)} sin coincidencia.")
        return no_match

    def resolver_no_matches(self, no_matches: list, diccionario: pd.DataFrame, fuente: str) -> list:
        nuevas_entradas = []
        self.append_log(f"Resolviendo no coincidencias para {fuente}...")

        for i, (nombre_raw, sugerencias) in enumerate(no_matches, start=1):
            self.append_log(f"  [{i}/{len(no_matches)}] {nombre_raw}")
            dialog = MismatchDialog(nombre_raw, sugerencias, self)
            if dialog.exec() != QDialog.DialogCode.Accepted:
                self.append_log("    Saltado por cancelación.")
                continue

            if dialog.selected_action == "assign":
                fila_elegida = dialog.selected_suggestion
                self.append_log(
                    f"    Asignado a [{fila_elegida[DICT_COL_ID]}] {fila_elegida[DICT_COL_NM]}"
                )
            elif dialog.selected_action == "add":
                nuevas_entradas.append(dialog.new_entry)
                self.append_log(
                    f"    Agregado nuevo: [{dialog.new_entry[DICT_COL_ID]}] {dialog.new_entry[DICT_COL_NM]}"
                )
            else:
                self.append_log("    Saltado.")

        return nuevas_entradas


if __name__ == "__main__":
    app = QApplication([])
    window = MainWindow()
    window.show()
    app.exec()
