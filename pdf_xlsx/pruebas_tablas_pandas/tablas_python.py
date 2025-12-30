from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers

# Crear workbook y seleccionar hoja activa
wb = Workbook()
ws = wb.active

# 1. FORMATO DE FONDO (FILL)
# Colores: 'FF0000' (rojo), '00FF00' (verde), '0000FF' (azul), 'FFFF00' (amarillo)
fill_rojo = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
fill_verde = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
fill_azul_claro = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

# 2. FORMATO DE FUENTE (FONT)
font_bold = Font(bold=True, color='FFFFFF', size=14)
font_italic = Font(italic=True, color='000000', size=12)
font_rojo = Font(color='FF0000', bold=True)

# 3. BORDES (BORDER)
border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

border_thick = Border(
    left=Side(style='thick'),
    right=Side(style='thick'),
    top=Side(style='thick'),
    bottom=Side(style='thick')
)

# 4. ALINEACIÓN (ALIGNMENT)
align_center = Alignment(horizontal='center', vertical='center')
align_right = Alignment(horizontal='right')
align_wrap = Alignment(wrap_text=True)

# 5. APLICAR FORMATOS A CELDAS
# Celda A1 - Título con fondo azul y texto blanco en negrita
ws['A1'] = 'TÍTULO PRINCIPAL'
ws['A1'].fill = fill_azul_claro
ws['A1'].font = font_bold
ws['A1'].alignment = align_center
ws['A1'].border = border_thin

# Celda B2 - Dato importante en rojo
ws['B2'] = 'ALERTA'
ws['B2'].font = font_rojo

# Celda C3 - Número con formato
ws['C3'] = 1234.5678
ws['C3'].number_format = '#,##0.00'  # Formato: 1,234.57

# Formato de moneda
ws['D4'] = 1500.75
ws['D4'].number_format = '"$"#,##0.00'  # $1,500.75

# Formato de porcentaje
ws['E5'] = 0.856
ws['E5'].number_format = '0.00%'  # 85.60%

# Formato de fecha
from datetime import datetime
ws['F6'] = datetime.now()
ws['F6'].number_format = 'dd/mm/yyyy hh:mm'

# 6. APLICAR FORMATO A RANGO DE CELDAS
for row in ws['A10:C15']:
    for cell in row:
        cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        cell.border = border_thin

# 7. AJUSTAR ANCHO DE COLUMNA
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 15

# 8. COMBINAR CELDAS
ws.merge_cells('A1:D1')

# Guardar archivo
wb.save('formato_celdas.xlsx')


