from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

def crear_formatos_predefinidos():
    """Retorna un diccionario con formatos predefinidos"""
    
    formatos = {
        'titulo': {
            'font': Font(bold=True, size=16, color='FFFFFF'),
            'fill': PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center')
        },
        
        'encabezado': {
            'font': Font(bold=True, size=12, color='000000'),
            'fill': PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        
        'moneda': {
            'number_format': '"$"#,##0.00',
            'alignment': Alignment(horizontal='right')
        },
        
        'porcentaje': {
            'number_format': '0.00%',
            'alignment': Alignment(horizontal='right')
        },
        
        'fecha': {
            'number_format': 'dd/mm/yyyy',
            'alignment': Alignment(horizontal='center')
        },
        
        'alerta': {
            'font': Font(color='FF0000', bold=True),
            'fill': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        },
        
        'exito': {
            'font': Font(color='00B050', bold=True),
            'fill': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        }
    }
    
    return formatos

# Uso de formatos predefinidos
wb = Workbook()
ws = wb.active
formatos = crear_formatos_predefinidos()

# Aplicar formato de título
for key, value in formatos['titulo'].items():
    setattr(ws['A1'], key, value)

# Guardar
wb.save('formatos_predefinidos.xlsx')