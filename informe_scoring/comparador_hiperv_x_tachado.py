"""
comparador_hipervinculos.py
Versión 1.2

Herramienta visual para:
  1. Seleccionar un archivo Excel que contiene hipervínculos.
  2. Seleccionar una carpeta donde se encuentran los PDF a comparar.
  3. Comparar los nombres de archivo de los hipervínculos con los PDF disponibles.
  4. Copiar los PDF coincidentes a una carpeta de destino elegida por el usuario.

Dependencias:
    pip install PyQt6 openpyxl
"""

import os
import shutil
import sys
from urllib.parse import unquote  # ← decodifica %20, %28, %29, etc.

import openpyxl
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QColor, QFont, QIcon
from PyQt6.QtWidgets import (
    QApplication,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QSizePolicy,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)


# ─────────────────────────────────────────────────────────────────────────────
#  WORKER — lógica pesada en hilo separado para no bloquear la UI
# ─────────────────────────────────────────────────────────────────────────────

class Worker(QThread):
    """
    Hilo de trabajo que:
      - Lee los hipervínculos del Excel.
      - Compara con los PDF de la carpeta origen.
      - Copia los coincidentes a la carpeta destino.

    Emite señales para comunicarse con la UI de forma segura.
    """

    # Señales que la UI escucha
    log        = pyqtSignal(str, str)   # (mensaje, nivel)  nivel: "ok"|"error"|"info"|"warn"
    progreso   = pyqtSignal(int)        # 0-100
    terminado  = pyqtSignal(int, int)   # (copiados, errores)

    def __init__(self, ruta_excel: str, carpeta_pdf: str, carpeta_destino: str):
        super().__init__()
        self.ruta_excel      = ruta_excel
        self.carpeta_pdf     = carpeta_pdf
        self.carpeta_destino = carpeta_destino

    # ------------------------------------------------------------------
    def run(self):
        """Punto de entrada del hilo."""
        copiados = 0
        errores  = 0

        # ── 1. Leer hipervínculos del Excel ────────────────────────────
        self.log.emit("Leyendo hipervínculos del archivo Excel…", "info")
        try:
            wb   = openpyxl.load_workbook(self.ruta_excel, data_only=True, read_only=False)
            hoja = wb.active
        except Exception as exc:
            self.log.emit(f"No se pudo abrir el Excel: {exc}", "error")
            self.terminado.emit(0, 1)
            return

        # ── Leer las reglas de formato condicional del workbook ─────────────
        # IMPORTANTE: openpyxl NO evalúa formato condicional en tiempo real.
        # Lo que hacemos es leer las REGLAS definidas y determinar qué rangos
        # tienen un dxf (DifferentialStyle) con font.strike == True.
        # Luego chequeamos si la coordenada de cada celda cae en esos rangos.
        #
        # Adicionalmente, también leemos el estilo DIRECTO de la celda
        # (celda.font.strike) por si el tachado fue aplicado manualmente.
        #
        # Ambas vías son necesarias porque Excel usa cualquiera de las dos.

        from openpyxl.utils import range_boundaries  # convierte "A1:C10" a (col1,row1,col2,row2)

        # Construir set de coordenadas que tienen tachado por formato condicional
        celdas_tachadas_cf: set[str] = set()
        for rango_str, reglas in hoja.conditional_formatting._cf_rules.items():
            for regla in reglas:
                # dxf es el DifferentialStyle de la regla; puede ser None
                if not regla.dxf:
                    continue
                if not regla.dxf.font:
                    continue
                if not regla.dxf.font.strike:
                    continue
                # Esta regla aplica tachado → marcar todas las celdas del rango
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(rango_str)
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            from openpyxl.utils import get_column_letter
                            coord = f"{get_column_letter(col)}{row}"
                            celdas_tachadas_cf.add(coord)
                except Exception:
                    pass  # rango malformado, ignorar

        # Recolectar hipervínculos que NO estén tachados (directamente ni por CF)
        nombres_hipervinculo: list[str] = []
        omitidos_tachado: int = 0

        for fila in hoja.iter_rows():
            for celda in fila:
                if not celda.hyperlink:
                    continue

                # ── Verificar tachado DIRECTO (estilo aplicado manualmente) ──
                tachado_directo = bool(celda.font and celda.font.strike)

                # ── Verificar tachado por FORMATO CONDICIONAL ─────────────
                tachado_cf = celda.coordinate in celdas_tachadas_cf

                if tachado_directo or tachado_cf:
                    target_omit = unquote(celda.hyperlink.target or "")
                    nombre_omit = os.path.basename(target_omit)
                    self.log.emit(f"[OMITIDO-TACHADO]  {nombre_omit}", "warn")
                    omitidos_tachado += 1
                    continue

                target = celda.hyperlink.target or ""
                # Ignorar URLs web y correos
                if target.startswith(("http://", "https://", "mailto:")):
                    continue
                # Decodificar %20, %28, %29, etc. que Excel agrega al editar
                target = unquote(target)
                nombre = os.path.basename(target)
                if nombre:
                    nombres_hipervinculo.append(nombre)

        if omitidos_tachado:
            self.log.emit(
                f"Hipervínculos omitidos por tachado: {omitidos_tachado}", "warn"
            )

        if not nombres_hipervinculo and omitidos_tachado == 0:
            self.log.emit("No se encontraron hipervínculos de archivos en el Excel.", "warn")
            self.terminado.emit(0, 0)
            return

        if not nombres_hipervinculo:
            self.log.emit("Todos los hipervínculos estaban tachados. Nada para copiar.", "warn")
            self.terminado.emit(0, 0)
            return

        self.log.emit(
            f"Hipervínculos activos (no tachados): {len(nombres_hipervinculo)}", "info"
        )

        # ── 2. Construir índice de PDF disponibles (nombre → ruta completa) ──
        self.log.emit("Indexando PDF en la carpeta origen…", "info")
        pdf_disponibles: dict[str, str] = {}
        try:
            for entrada in os.scandir(self.carpeta_pdf):
                if entrada.is_file() and entrada.name.lower().endswith(".pdf"):
                    # Guardamos en minúsculas para comparación insensible a mayúsculas
                    pdf_disponibles[entrada.name.lower()] = entrada.path
        except Exception as exc:
            self.log.emit(f"Error al leer la carpeta de PDF: {exc}", "error")
            self.terminado.emit(0, 1)
            return

        self.log.emit(f"PDF encontrados en carpeta origen: {len(pdf_disponibles)}", "info")

        # ── 3. Crear carpeta destino si no existe ──────────────────────
        try:
            os.makedirs(self.carpeta_destino, exist_ok=True)
        except Exception as exc:
            self.log.emit(f"No se pudo crear la carpeta destino: {exc}", "error")
            self.terminado.emit(0, 1)
            return

        # ── 4. Comparar y copiar ───────────────────────────────────────
        total = len(nombres_hipervinculo)
        for idx, nombre in enumerate(nombres_hipervinculo, start=1):
            # Actualizar barra de progreso
            self.progreso.emit(int(idx / total * 100))

            clave = nombre.lower()
            if clave in pdf_disponibles:
                ruta_origen  = pdf_disponibles[clave]
                ruta_destino = os.path.join(self.carpeta_destino, nombre)
                try:
                    shutil.copy2(ruta_origen, ruta_destino)
                    self.log.emit(f"[COPIADO]  {nombre}", "ok")
                    copiados += 1
                except Exception as exc:
                    self.log.emit(f"[ERROR AL COPIAR]  {nombre} → {exc}", "error")
                    errores += 1
            else:
                self.log.emit(f"[NO ENCONTRADO]  {nombre}", "error")
                errores += 1

        self.progreso.emit(100)
        self.terminado.emit(copiados, errores)


# ─────────────────────────────────────────────────────────────────────────────
#  VENTANA PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

class VentanaPrincipal(QMainWindow):

    # Paleta de colores de la aplicación
    COLOR_FONDO       = "#1C1C2E"   # azul noche profundo
    COLOR_PANEL       = "#252538"   # panel ligeramente más claro
    COLOR_ACENTO      = "#7C5CBF"   # violeta corporativo
    COLOR_ACENTO2     = "#A78BFA"   # violeta claro para hover/texto
    COLOR_TEXTO       = "#E2E8F0"   # blanco azulado suave
    COLOR_TEXTO_SEC   = "#94A3B8"   # gris azulado
    COLOR_OK          = "#4ADE80"   # verde éxito
    COLOR_ERROR       = "#F87171"   # rojo error
    COLOR_WARN        = "#FBBF24"   # amarillo aviso
    COLOR_BORDE       = "#3D3D5C"   # borde sutil

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Comparador de Hipervínculos PDF  v1.2")
        self.setMinimumSize(820, 640)
        self._worker: Worker | None = None
        self._init_ui()
        self._aplicar_estilos()

    # ──────────────────────────────────────────────────────────────────
    #  CONSTRUCCIÓN DE LA UI
    # ──────────────────────────────────────────────────────────────────

    def _init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        raiz = QVBoxLayout(central)
        raiz.setContentsMargins(24, 24, 24, 24)
        raiz.setSpacing(16)

        # ── Título ─────────────────────────────────────────────────────
        titulo = QLabel("Comparador de Hipervínculos  →  PDF")
        titulo.setObjectName("titulo")
        raiz.addWidget(titulo)

        subtitulo = QLabel(
            "Seleccioná el Excel con hipervínculos, la carpeta de PDFs y "
            "dónde guardar los coincidentes."
        )
        subtitulo.setObjectName("subtitulo")
        subtitulo.setWordWrap(True)
        raiz.addWidget(subtitulo)

        raiz.addWidget(self._separador())

        # ── Sección de rutas ───────────────────────────────────────────
        raiz.addWidget(self._bloque_ruta(
            etiqueta="Archivo Excel con hipervínculos",
            placeholder="Ej: D:/datos/contenido_archivos.xlsx",
            boton_texto="Examinar…",
            attr_linea="campo_excel",
            slot=self._elegir_excel,
        ))

        raiz.addWidget(self._bloque_ruta(
            etiqueta="Carpeta con los PDF a comparar",
            placeholder="Ej: D:/pdfs/originales",
            boton_texto="Examinar…",
            attr_linea="campo_carpeta_pdf",
            slot=self._elegir_carpeta_pdf,
        ))

        raiz.addWidget(self._bloque_ruta(
            etiqueta="Carpeta destino (donde se copiarán los coincidentes)",
            placeholder="Ej: D:/pdfs/coincidentes",
            boton_texto="Examinar…",
            attr_linea="campo_destino",
            slot=self._elegir_destino,
        ))

        raiz.addWidget(self._separador())

        # ── Botón principal ────────────────────────────────────────────
        self.btn_iniciar = QPushButton("▶  Iniciar comparación")
        self.btn_iniciar.setObjectName("btnPrimario")
        self.btn_iniciar.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_iniciar.clicked.connect(self._iniciar)
        raiz.addWidget(self.btn_iniciar)

        # ── Barra de progreso ──────────────────────────────────────────
        self.progreso = QProgressBar()
        self.progreso.setValue(0)
        self.progreso.setTextVisible(True)
        self.progreso.setFormat("%p%")
        self.progreso.setFixedHeight(10)
        raiz.addWidget(self.progreso)

        # ── Log de resultados ──────────────────────────────────────────
        lbl_log = QLabel("Registro de operaciones")
        lbl_log.setObjectName("labelSeccion")
        raiz.addWidget(lbl_log)

        self.log_area = QTextEdit()
        self.log_area.setReadOnly(True)
        self.log_area.setObjectName("logArea")
        self.log_area.setFont(QFont("Consolas", 9))
        raiz.addWidget(self.log_area)

        # ── Botón limpiar ──────────────────────────────────────────────
        btn_limpiar = QPushButton("Limpiar registro")
        btn_limpiar.setObjectName("btnSecundario")
        btn_limpiar.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_limpiar.clicked.connect(self.log_area.clear)
        btn_limpiar.setSizePolicy(
            QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed
        )
        raiz.addWidget(btn_limpiar, alignment=Qt.AlignmentFlag.AlignRight)

    # ──────────────────────────────────────────────────────────────────
    #  HELPERS DE UI
    # ──────────────────────────────────────────────────────────────────

    def _bloque_ruta(
        self,
        etiqueta: str,
        placeholder: str,
        boton_texto: str,
        attr_linea: str,
        slot,
    ) -> QWidget:
        """Genera un bloque: Etiqueta + QLineEdit + Botón."""
        contenedor = QWidget()
        vbox = QVBoxLayout(contenedor)
        vbox.setContentsMargins(0, 0, 0, 0)
        vbox.setSpacing(6)

        lbl = QLabel(etiqueta)
        lbl.setObjectName("labelCampo")
        vbox.addWidget(lbl)

        fila = QHBoxLayout()
        fila.setSpacing(8)

        campo = QLineEdit()
        campo.setPlaceholderText(placeholder)
        campo.setObjectName("campoCampo")
        fila.addWidget(campo)

        boton = QPushButton(boton_texto)
        boton.setObjectName("btnExaminar")
        boton.setCursor(Qt.CursorShape.PointingHandCursor)
        boton.setFixedWidth(110)
        boton.clicked.connect(slot)
        fila.addWidget(boton)

        vbox.addLayout(fila)
        # Guardar referencia al campo para acceder luego
        setattr(self, attr_linea, campo)
        return contenedor

    @staticmethod
    def _separador() -> QFrame:
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setObjectName("separador")
        return sep

    # ──────────────────────────────────────────────────────────────────
    #  ESTILOS (QSS — equivalente a CSS para Qt)
    # ──────────────────────────────────────────────────────────────────

    def _aplicar_estilos(self):
        self.setStyleSheet(f"""
        /* ── Ventana y widget base ── */
        QMainWindow, QWidget {{
            background-color: {self.COLOR_FONDO};
            color: {self.COLOR_TEXTO};
            font-family: 'Segoe UI', Arial, sans-serif;
            font-size: 13px;
        }}

        /* ── Título principal ── */
        QLabel#titulo {{
            font-size: 20px;
            font-weight: 700;
            color: {self.COLOR_ACENTO2};
            letter-spacing: 0.5px;
        }}

        /* ── Subtítulo ── */
        QLabel#subtitulo {{
            font-size: 12px;
            color: {self.COLOR_TEXTO_SEC};
        }}

        /* ── Etiqueta de sección ── */
        QLabel#labelSeccion {{
            font-size: 12px;
            font-weight: 600;
            color: {self.COLOR_TEXTO_SEC};
            text-transform: uppercase;
            letter-spacing: 1px;
        }}

        /* ── Etiqueta de campo ── */
        QLabel#labelCampo {{
            font-size: 12px;
            font-weight: 600;
            color: {self.COLOR_ACENTO2};
        }}

        /* ── Inputs de texto ── */
        QLineEdit#campoCampo {{
            background-color: {self.COLOR_PANEL};
            border: 1px solid {self.COLOR_BORDE};
            border-radius: 6px;
            padding: 7px 10px;
            color: {self.COLOR_TEXTO};
        }}
        QLineEdit#campoCampo:focus {{
            border: 1px solid {self.COLOR_ACENTO};
        }}

        /* ── Botón "Examinar" ── */
        QPushButton#btnExaminar {{
            background-color: {self.COLOR_PANEL};
            border: 1px solid {self.COLOR_BORDE};
            border-radius: 6px;
            padding: 7px 12px;
            color: {self.COLOR_TEXTO_SEC};
            font-weight: 600;
        }}
        QPushButton#btnExaminar:hover {{
            border-color: {self.COLOR_ACENTO};
            color: {self.COLOR_ACENTO2};
        }}

        /* ── Botón primario ── */
        QPushButton#btnPrimario {{
            background-color: {self.COLOR_ACENTO};
            border: none;
            border-radius: 8px;
            padding: 10px 20px;
            color: #FFFFFF;
            font-size: 14px;
            font-weight: 700;
        }}
        QPushButton#btnPrimario:hover {{
            background-color: {self.COLOR_ACENTO2};
            color: {self.COLOR_FONDO};
        }}
        QPushButton#btnPrimario:disabled {{
            background-color: {self.COLOR_BORDE};
            color: {self.COLOR_TEXTO_SEC};
        }}

        /* ── Botón secundario ── */
        QPushButton#btnSecundario {{
            background-color: transparent;
            border: 1px solid {self.COLOR_BORDE};
            border-radius: 6px;
            padding: 5px 14px;
            color: {self.COLOR_TEXTO_SEC};
            font-size: 11px;
        }}
        QPushButton#btnSecundario:hover {{
            border-color: {self.COLOR_ACENTO};
            color: {self.COLOR_ACENTO2};
        }}

        /* ── Área de log ── */
        QTextEdit#logArea {{
            background-color: {self.COLOR_PANEL};
            border: 1px solid {self.COLOR_BORDE};
            border-radius: 8px;
            padding: 10px;
            color: {self.COLOR_TEXTO};
            font-family: 'Consolas', monospace;
        }}

        /* ── Barra de progreso ── */
        QProgressBar {{
            background-color: {self.COLOR_PANEL};
            border: 1px solid {self.COLOR_BORDE};
            border-radius: 5px;
            text-align: center;
            color: {self.COLOR_TEXTO};
        }}
        QProgressBar::chunk {{
            background-color: {self.COLOR_ACENTO};
            border-radius: 4px;
        }}

        /* ── Separador ── */
        QFrame#separador {{
            color: {self.COLOR_BORDE};
        }}
        """)

    # ──────────────────────────────────────────────────────────────────
    #  SLOTS — selección de rutas
    # ──────────────────────────────────────────────────────────────────

    def _elegir_excel(self):
        ruta, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar archivo Excel",
            "",
            "Archivos Excel (*.xlsx *.xlsm *.xls)",
        )
        if ruta:
            self.campo_excel.setText(ruta)

    def _elegir_carpeta_pdf(self):
        ruta = QFileDialog.getExistingDirectory(
            self, "Seleccionar carpeta con PDFs"
        )
        if ruta:
            self.campo_carpeta_pdf.setText(ruta)

    def _elegir_destino(self):
        ruta = QFileDialog.getExistingDirectory(
            self, "Seleccionar carpeta destino"
        )
        if ruta:
            self.campo_destino.setText(ruta)

    # ──────────────────────────────────────────────────────────────────
    #  SLOT — iniciar proceso
    # ──────────────────────────────────────────────────────────────────

    def _iniciar(self):
        # ── Validaciones básicas ───────────────────────────────────────
        excel   = self.campo_excel.text().strip()
        carpeta = self.campo_carpeta_pdf.text().strip()
        destino = self.campo_destino.text().strip()

        if not excel:
            self._alerta("Falta el archivo Excel", "Por favor seleccioná el archivo Excel.")
            return
        if not os.path.isfile(excel):
            self._alerta("Archivo no encontrado", f"El archivo Excel no existe:\n{excel}")
            return
        if not carpeta:
            self._alerta("Falta la carpeta de PDFs", "Por favor seleccioná la carpeta con los PDFs.")
            return
        if not os.path.isdir(carpeta):
            self._alerta("Carpeta no encontrada", f"La carpeta de PDFs no existe:\n{carpeta}")
            return
        if not destino:
            self._alerta("Falta la carpeta destino", "Por favor seleccioná dónde guardar los PDF coincidentes.")
            return

        # ── Lanzar worker ──────────────────────────────────────────────
        self.btn_iniciar.setEnabled(False)
        self.progreso.setValue(0)
        self.log_area.clear()
        self._escribir_log("═" * 60, "info")
        self._escribir_log("  INICIO DEL PROCESO", "info")
        self._escribir_log("═" * 60, "info")

        self._worker = Worker(excel, carpeta, destino)
        self._worker.log.connect(self._escribir_log)
        self._worker.progreso.connect(self.progreso.setValue)
        self._worker.terminado.connect(self._al_terminar)
        self._worker.start()

    # ──────────────────────────────────────────────────────────────────
    #  SLOT — proceso terminado
    # ──────────────────────────────────────────────────────────────────

    def _al_terminar(self, copiados: int, errores: int):
        self.btn_iniciar.setEnabled(True)
        self._escribir_log("═" * 60, "info")
        self._escribir_log(
            f"  FINALIZADO  |  Copiados: {copiados}   No encontrados / errores: {errores}",
            "ok" if errores == 0 else "warn",
        )
        self._escribir_log("═" * 60, "info")

        # Diálogo de resumen
        icono = QMessageBox.Icon.Information if errores == 0 else QMessageBox.Icon.Warning
        msg = QMessageBox(self)
        msg.setIcon(icono)
        msg.setWindowTitle("Proceso finalizado")
        msg.setText(
            f"<b>Archivos copiados:</b> {copiados}<br>"
            f"<b>No encontrados / errores:</b> {errores}<br>"
            f"<b>Omitidos por tachado:</b> ver registro"
        )
        msg.exec()

    # ──────────────────────────────────────────────────────────────────
    #  HELPERS
    # ──────────────────────────────────────────────────────────────────

    def _escribir_log(self, mensaje: str, nivel: str):
        """Agrega una línea coloreada al área de log."""
        colores = {
            "ok":    self.COLOR_OK,
            "error": self.COLOR_ERROR,
            "warn":  self.COLOR_WARN,
            "info":  self.COLOR_TEXTO_SEC,
        }
        color = colores.get(nivel, self.COLOR_TEXTO)
        self.log_area.append(
            f'<span style="color:{color};">{mensaje}</span>'
        )
        # Auto-scroll al fondo
        sb = self.log_area.verticalScrollBar()
        sb.setValue(sb.maximum())

    @staticmethod
    def _alerta(titulo: str, mensaje: str):
        dlg = QMessageBox()
        dlg.setWindowTitle(titulo)
        dlg.setText(mensaje)
        dlg.setIcon(QMessageBox.Icon.Warning)
        dlg.exec()


# ─────────────────────────────────────────────────────────────────────────────
#  PUNTO DE ENTRADA
# ─────────────────────────────────────────────────────────────────────────────

def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")          # Base neutral → los estilos QSS toman el control
    ventana = VentanaPrincipal()
    ventana.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
