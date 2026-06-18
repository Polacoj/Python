"""
Analizador de Personal con interfaz PyQt6.

Este script carga los tres archivos requeridos, valida las hojas y columnas,
reconciliando los nombres de OPERADOR con el Diccionario Maestro.

Flujo principal:
  1. Selecciona Diccionario Maestro, Fiscalización Diaria/Mensual y Control Diario Mes.
  2. Carga y selecciona hoja si hay múltiples hojas.
  3. Valida columnas requeridas.
  4. Reconciliación de Fiscalización Diaria/Mensual y Control Diario.
  5. Guarda archivos procesados y/o diccionario actualizado.
"""

import os
from pathlib import Path

import pandas as pd
from PyQt6.QtWidgets import (
    QApplication,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

# Columnas esperadas
COL_ID_DICCIONARIO = "LP"
COL_NOMBRE_DICCIONARIO = "APELLIDO Y NOMBRE"
COLUMNAS_FISCALIZACION_DIARIA = ["OFICIAL"]


def normalizar_columnas(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(col).strip().upper() for col in df.columns]
    return df


def validar_columnas(
    df: pd.DataFrame, columnas_requeridas: list[str]
) -> tuple[bool, str]:
    present = set(df.columns)
    faltantes = [col for col in columnas_requeridas if col not in present]
    if faltantes:
        mensaje = (
            f"Columnas faltantes: {faltantes}. Columnas encontradas: {sorted(present)}"
        )
        return False, mensaje
    return True, ""


def normalizar_nombre(valor) -> str:
    if pd.isna(valor) or str(valor).strip() == "":
        return ""
    return " ".join(str(valor).upper().split())


def construir_mapa_nombres(df_diccionario: pd.DataFrame) -> dict[str, str]:
    mapa: dict[str, str] = {}
    for _, fila in df_diccionario.iterrows():
        nombre_raw = fila[COL_NOMBRE_DICCIONARIO]
        id_raw = fila[COL_ID_DICCIONARIO]
        if pd.isna(nombre_raw) or str(nombre_raw).strip() == "":
            continue
        nombre_norm = normalizar_nombre(nombre_raw)
        id_str = "" if pd.isna(id_raw) else str(id_raw).strip()
        mapa[nombre_norm] = id_str
    return mapa


def leer_archivo_excel(ruta: Path) -> dict[str, pd.DataFrame]:
    return pd.read_excel(ruta, sheet_name=None, dtype=str)


class SheetSelectionDialog(QDialog):
    def __init__(self, títulos: list[str], parent=None):
        super().__init__(parent)
        self.setWindowTitle("Seleccionar hoja")
        self.selected_sheet = None
        self.setup_ui(títulos)

    def setup_ui(self, títulos: list[str]):
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("El archivo tiene varias hojas. Seleccioná una:"))
        self.combo = QComboBox()
        self.combo.addItems(títulos)
        layout.addWidget(self.combo)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def accept(self) -> None:
        self.selected_sheet = self.combo.currentText()
        super().accept()



def guardar_diccionario_con_bak(df_actualizado, ruta_original):
    import shutil
    from openpyxl import load_workbook
    ruta_bak = ruta_original.parent / "ANTIGUO_DICCIONARIO.xlsx"
    shutil.copy2(ruta_original, ruta_bak)
    wb = load_workbook(ruta_original)
    nombre_sheet1 = next(
        (s for s in wb.sheetnames if s.strip().upper() == "SHEET1"), None
    )
    if nombre_sheet1 is None:
        ws = wb.create_sheet("Sheet1")
    else:
        ws = wb[nombre_sheet1]
        ws.delete_rows(1, ws.max_row)
    for col_idx, enc in enumerate(df_actualizado.columns, start=1):
        ws.cell(row=1, column=col_idx, value=enc)
    for fila_idx, fila in enumerate(df_actualizado.itertuples(index=False), start=2):
        for col_idx, valor in enumerate(fila, start=1):
            import pandas as pd
            ws.cell(row=fila_idx, column=col_idx,
                    value="" if pd.isna(valor) else valor)
    wb.save(ruta_original)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Analizador de Personal")
        self.resize(900, 700)
        # Rutas de los _PROCESADO y del diccionario: se asignan al
        # finalizar el proceso directo con éxito y las usa on_revertir().
        self._ruta_trabajo_procesado: Path | None = None
        self._ruta_diccionario: Path | None = None
        self.setup_ui()

    def setup_ui(self):
        central = QWidget()
        layout = QVBoxLayout(central)

        self.dict_path = QLineEdit()
        self.trabajo_path = QLineEdit()

        dict_btn = QPushButton("Seleccionar")
        trabajo_btn = QPushButton("Seleccionar")
        dict_btn.clicked.connect(lambda: self.select_file(self.dict_path))
        trabajo_btn.clicked.connect(lambda: self.select_file(self.trabajo_path))

        layout.addLayout(
            self.file_row("Diccionario Maestro OFICIALES:", self.dict_path, dict_btn)
        )

        layout.addLayout(
            self.file_row(
                "Archivo Fiscalizacion diaria/mensual:", self.trabajo_path, trabajo_btn
            )
        )

        self.start_button = QPushButton("Validar y reconciliar")
        self.start_button.clicked.connect(self.on_start)
        self.start_button.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border-radius: 8px;
                font-size: 16px;
                padding: 10px;
            }
            QPushButton:hover {
                background-color: #0b7dda;
            }
        """)
        layout.addWidget(self.start_button)

        # Botón proceso inverso: deshabilitado hasta que ambos módulos sean 100%
        self.reverse_button = QPushButton("↩  Revertir IDs → Nombres (hoja Validacion)")
        self.reverse_button.setEnabled(False)
        self.reverse_button.clicked.connect(self.on_revertir)
        self.reverse_button.setStyleSheet("""
            QPushButton {
                background-color: #9E9E9E;
                color: white;
                border-radius: 8px;
                font-size: 16px;
                padding: 10px;
            }
            QPushButton:enabled {
                background-color: #4CAF50;
            }
            QPushButton:enabled:hover {
                background-color: #388E3C;
            }
        """)
        layout.addWidget(self.reverse_button)

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        layout.addWidget(QLabel("Registro de ejecución:"))
        self.log_text.setStyleSheet("""
            QTextEdit {
                border-radius: 10px;
                font-size: 16px;
                padding: 10px;
            }
        """)
        layout.addWidget(self.log_text)

        self.setCentralWidget(central)

        self.exit_button = QPushButton("Salir")
        self.exit_button.clicked.connect(self.close)
        self.exit_button.setStyleSheet("""
            QPushButton {
                background-color: #eb7f82;
                color: white;
                border-radius: 8px;
                font-size: 16px;
                padding: 10px;
            }
            QPushButton:hover {
                background-color: #c24a4e;
            }
        """)
        layout.addWidget(self.exit_button)

    def file_row(
        self, label_text: str, line_edit: QLineEdit, button: QPushButton
    ) -> QHBoxLayout:
        row = QHBoxLayout()
        row.addWidget(QLabel(label_text))
        row.addWidget(line_edit)
        row.addWidget(button)
        return row

    def select_file(self, line_edit: QLineEdit) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar archivo Excel",
            os.getcwd(),
            "Archivos Excel (*.xlsx *.xlsm);Todos los archivos (*)",
        )
        if path:
            line_edit.setText(path)

    def append_log(self, mensaje: str) -> None:
        self.log_text.append(mensaje)
        QApplication.processEvents()

    def on_start(self) -> None:
        self.start_button.setEnabled(False)
        self.log_text.clear()
        try:
            self.run_process()
        except Exception as e:
            self.append_log(f"ERROR: {e}")
            QMessageBox.critical(self, "Error", str(e))
        finally:
            self.start_button.setEnabled(True)

    def run_process(self) -> None:
        ruta_dict = self.validate_path(
            self.dict_path.text().strip(), "Diccionario Maestro"
        )
        ruta_trabajo = self.validate_path(
            self.trabajo_path.text().strip(), "Fiscalización diaria/mensual"
        )

        dfs_dict = self.load_excel_with_sheet_selection(
            ruta_dict, "Diccionario Maestro"
        )
        df_dict = normalizar_columnas(dfs_dict)

        dfs_trabajo = self.load_excel_with_sheet_selection(
            ruta_trabajo, "Fiscalización diaria/mensual"
        )
        df_trabajo = normalizar_columnas(dfs_trabajo)
        valido_trabajo, mensaje_trabajo = validar_columnas(
            df_trabajo, COLUMNAS_FISCALIZACION_DIARIA
        )
        if not valido_trabajo:
            raise RuntimeError(f"Fiscalización Diaria/Mensual: {mensaje_trabajo}")

        self.append_log("Archivos cargados y validados correctamente.")
        self.append_log(f"Diccionario: {ruta_dict.name}\n")
        self.append_log(f"Fiscalización Diaria/Mensual: {ruta_trabajo.name}\n")

        df_trabajo_mod, _ = self.reconciliar_trabajo_diario(
            df_trabajo, df_dict, ruta_trabajo, ruta_dict
        )
        self.append_log("Módulo completado.\n")

        self._ruta_trabajo_procesado = ruta_trabajo.parent / (
            ruta_trabajo.stem + "_PROCESADO" + ruta_trabajo.suffix
        )
        self._ruta_diccionario = ruta_dict
        self.reverse_button.setEnabled(True)
        self.append_log(
            "\n✔ Proceso inverso disponible. Presioná '↩ Revertir IDs → Nombres'."
        )

        QMessageBox.information(self, "Finalizado", "Proceso completado correctamente.")

    def validate_path(self, ruta_text: str, etiqueta: str) -> Path:
        if ruta_text == "":
            raise RuntimeError(f"Debés seleccionar el archivo: {etiqueta}")
        ruta = Path(ruta_text)
        if not ruta.exists():
            raise RuntimeError(f"No se encontró el archivo {etiqueta}: {ruta}")
        return ruta

    def load_excel_with_sheet_selection(
        self, ruta: Path, etiqueta: str
    ) -> pd.DataFrame:
        self.append_log(f"Cargando '{ruta.name}'...")
        hojas = leer_archivo_excel(ruta)
        if len(hojas) == 0:
            raise RuntimeError(f"El archivo {etiqueta} no tiene hojas válidas.")
        if len(hojas) == 1:
            nombre = next(iter(hojas))
            self.append_log(f"Hoja seleccionada automáticamente: {nombre}")
            return hojas[nombre]

        dialog = SheetSelectionDialog(list(hojas.keys()), self)
        if dialog.exec() != QDialog.DialogCode.Accepted or not dialog.selected_sheet:
            raise RuntimeError(f"Se canceló la selección de hoja para {etiqueta}.")
        self.append_log(f"Hoja seleccionada: {dialog.selected_sheet}")
        return hojas[dialog.selected_sheet]

    def reconciliar_trabajo_diario(
        self,
        df_trabajo: pd.DataFrame,
        df_diccionario: pd.DataFrame,
        ruta_trabajo: Path,
        ruta_diccionario: Path,
    ) -> tuple[pd.DataFrame | None, pd.DataFrame]:
        self.append_log("Iniciando reconciliación de fiscalización diaria/mensual...")

        if (
            COL_ID_DICCIONARIO not in df_diccionario.columns
            or COL_NOMBRE_DICCIONARIO not in df_diccionario.columns
        ):
            raise RuntimeError(
                f"El Diccionario Maestro debe contener las columnas '{COL_ID_DICCIONARIO}' y '{COL_NOMBRE_DICCIONARIO}'."
            )

        mapa_nombres = construir_mapa_nombres(df_diccionario)
        reemplazados = 0
        no_encontrados = 0
        vacias = 0
        nombres_nuevos: dict[str, list[int]] = {}

        df_trabajo_mod = df_trabajo.copy()
        resultados: list[tuple[int, str, str | None]] = []

        if "OFICIAL" not in df_trabajo_mod.columns:
            raise RuntimeError(
                "La hoja de fiscalización diaria/mensual no contiene la columna 'OFICIAL'."
            )

        for idx, valor_celda in df_trabajo_mod["OFICIAL"].items():
            idx = int(idx)  # asegurar que el índice es entero
            nombre_norm = normalizar_nombre(valor_celda)
            if nombre_norm == "":
                vacias += 1
                resultados.append((idx, str(valor_celda), "VACIO"))
                continue
            if nombre_norm in mapa_nombres:
                id_encontrado = mapa_nombres[nombre_norm]
                reemplazados += 1
                resultados.append((idx, str(valor_celda), id_encontrado))
            else:
                no_encontrados += 1
                resultados.append((idx, str(valor_celda), None))
                nombres_nuevos.setdefault(nombre_norm, []).append(idx + 1)

        celdas_evaluadas = reemplazados + no_encontrados
        porcentaje = (
            100.0 if celdas_evaluadas == 0 else (reemplazados / celdas_evaluadas) * 100
        )
        self.append_log(
            f"Fiscalización Diaria/Mensual: {celdas_evaluadas} celdas evaluadas, {reemplazados} coincidencias, {no_encontrados} no encontradas, {porcentaje:.1f}%."
        )

        if porcentaje < 100.0:
            self.append_log(
                "Coincidencia incompleta. Actualizando diccionario original con los faltantes..."
            )
            nuevas_filas = []
            for nombre_norm, filas in nombres_nuevos.items():
                fila_nueva = {col: "" for col in df_diccionario.columns}
                fila_nueva[COL_NOMBRE_DICCIONARIO] = nombre_norm
                nuevas_filas.append(fila_nueva)
                self.append_log(f"  Nombre nuevo: '{nombre_norm}' en filas {filas}")
            df_diccionario_act = pd.concat(
                [df_diccionario, pd.DataFrame(nuevas_filas)], ignore_index=True
            )
            guardar_diccionario_con_bak(df_diccionario_act, ruta_diccionario)
            self.append_log("  ✔ Backup guardado: ANTIGUO_DICCIONARIO.xlsx")
            self.append_log(f"  ✔ Diccionario original actualizado: {ruta_diccionario.name}")
            raise RuntimeError(
                f"Coincidencia incompleta ({porcentaje:.1f}%). "
                f"Se actualizó '{ruta_diccionario.name}' con los nombres faltantes. "
                f"Completá la columna ID y volvé a correr."
            )

        for idx, _valor_original, id_valor in resultados:
            if id_valor is None or id_valor == "VACIO":
                continue
            df_trabajo_mod.at[idx, "OFICIAL"] = (
                id_valor  # ------------------------------------------
            )

        ruta_trabajo_salida = ruta_trabajo.parent / (
            ruta_trabajo.stem + "_PROCESADO" + ruta_trabajo.suffix
        )
        df_trabajo_mod.to_excel(ruta_trabajo_salida, index=False)
        self.append_log(
            f"Fiscalización Diaria/mensual procesada guardada en: {ruta_trabajo_salida.name}"
        )
        return df_trabajo_mod, df_diccionario

    def on_revertir(self) -> None:
        """Dispara el proceso inverso y maneja errores con feedback visual."""
        self.reverse_button.setEnabled(False)
        try:
            self.revertir_ids_a_nombres()
        except Exception as e:
            self.append_log(f"ERROR en proceso inverso: {e}")
            QMessageBox.critical(self, "Error", str(e))
        finally:
            self.reverse_button.setEnabled(True)

    def revertir_ids_a_nombres(self) -> None:
        """
        Proceso inverso: reemplaza los IDs (LP/DNI) en los archivos _PROCESADO
        por los nombres correspondientes según la hoja 'VALIDACION' del
        Diccionario Maestro.

        Flujo:
          1. Leer la hoja 'VALIDACION' del Diccionario Maestro.
          2. Construir mapa inverso { ID → APELLIDO Y NOMBRE }.
          3. Para cada columna de personal en ambos archivos _PROCESADO:
               - Buscar el ID en el mapa inverso.
               - Reemplazar por el nombre si hay coincidencia.
               - Acumular IDs sin coincidencia (no debería haber, pero se reportan).
          4. Guardar como _REVERTIDO.xlsx (preserva los _PROCESADO intactos).
        """
        self.append_log("\n" + "═" * 50)
        self.append_log("  PROCESO INVERSO: IDs → NOMBRES (hoja VALIDACION)")
        self.append_log("═" * 50)

        # ── Verificar que el proceso directo se haya ejecutado ────
        if not self._ruta_trabajo_procesado:
            raise RuntimeError(
                "No se encontró el archivo _PROCESADO. "
                "Ejecutá primero el proceso de validación y reconciliación."
            )
        if not self._ruta_trabajo_procesado.exists():
            raise RuntimeError(
                f"No se encontró el archivo: {self._ruta_trabajo_procesado.name}"
            )

        # ── Paso 1: Leer hoja VALIDACION del diccionario ─────────
        self.append_log(
            f"\nLeyendo hoja 'VALIDACION' de: {self._ruta_diccionario.name}..."
        )
        todas_hojas = leer_archivo_excel(self._ruta_diccionario)

        # Buscar la hoja ignorando mayúsculas/minúsculas
        hoja_val = next(
            (
                nombre
                for nombre in todas_hojas
                if nombre.strip().upper() == "VALIDACION"
            ),
            None,
        )
        if hoja_val is None:
            hojas_disponibles = list(todas_hojas.keys())
            raise RuntimeError(
                f"No se encontró la hoja 'VALIDACION' en el Diccionario Maestro.\n"
                f"Hojas disponibles: {hojas_disponibles}"
            )

        df_val = normalizar_columnas(todas_hojas[hoja_val])
        self.append_log(f"  Hoja encontrada: '{hoja_val}' — {len(df_val)} filas.")

        # Verificar columnas requeridas en la hoja de validación
        for col_req in [COL_ID_DICCIONARIO, COL_NOMBRE_DICCIONARIO]:
            if col_req not in df_val.columns:
                raise RuntimeError(
                    f"La hoja 'VALIDACION' no contiene la columna '{col_req}'.\n"
                    f"Columnas encontradas: {list(df_val.columns)}"
                )

        # ── Paso 2: Construir mapa inverso { ID → NOMBRE } ───────
        # A diferencia del proceso directo (nombre → ID),
        # aquí la clave es el ID y el valor es el nombre.
        mapa_inverso: dict[str, str] = {}
        for _, fila in df_val.iterrows():
            id_raw = fila[COL_ID_DICCIONARIO]
            nombre_raw = fila[COL_NOMBRE_DICCIONARIO]
            if pd.isna(id_raw) or str(id_raw).strip() == "":
                continue  # filas sin ID no aportan al mapa
            id_str = str(id_raw).strip()
            nombre_str = str(nombre_raw).strip() if not pd.isna(nombre_raw) else ""
            mapa_inverso[id_str] = nombre_str

        self.append_log(f"  Mapa inverso construido: {len(mapa_inverso)} entradas.")

        # ── Paso 3: Revertir el archivo _PROCESADO ───────────────
        self._revertir_archivo(
            ruta_procesado=self._ruta_trabajo_procesado,
            columnas=["OFICIAL"],
            mapa_inverso=mapa_inverso,
            etiqueta="Fiscalización Diaria/Mensual",
        )

        self.append_log("\n" + "═" * 50)
        self.append_log("  ✔ PROCESO INVERSO COMPLETADO")
        self.append_log("═" * 50)
        QMessageBox.information(
            self, "Proceso inverso", "Reversión completada correctamente."
        )

    def _revertir_archivo(
        self,
        ruta_procesado: Path,
        columnas: list[str],
        mapa_inverso: dict[str, str],
        etiqueta: str,
    ) -> None:
        """
        Revierte un archivo _PROCESADO: reemplaza IDs por nombres usando
        el mapa inverso y guarda el resultado como _REVERTIDO.xlsx.

        Parámetros:
          ruta_procesado (Path):       Ruta al archivo _PROCESADO.
          columnas (list[str]):        Columnas a procesar en ese archivo.
          mapa_inverso (dict):         { ID_str → nombre_str }
          etiqueta (str):              Nombre descriptivo para el log.
        """
        self.append_log(f"\nProcesando {etiqueta}: {ruta_procesado.name}...")

        # Leer el _PROCESADO (tiene una sola hoja, tomamos la primera)
        hojas = leer_archivo_excel(ruta_procesado)
        nombre_hoja = next(iter(hojas))
        df = normalizar_columnas(hojas[nombre_hoja])

        reemplazados = 0
        sin_match = 0
        vacias = 0
        ids_sin_match: list[str] = []

        for col in columnas:
            if col not in df.columns:
                self.append_log(
                    f"  ⚠ Columna '{col}' no encontrada en {ruta_procesado.name}, se omite."
                )
                continue

            for idx, valor_celda in df[col].items():
                valor_str = str(valor_celda).strip()

                # Celda vacía o NaN → ignorar
                if (
                    pd.isna(valor_celda)
                    or valor_str == ""
                    or valor_str.lower() == "nan"
                ):
                    vacias += 1
                    continue

                # Buscar el ID en el mapa inverso
                if valor_str in mapa_inverso:
                    nombre_revertido = mapa_inverso[valor_str]
                    df.at[idx, col] = nombre_revertido
                    reemplazados += 1
                else:
                    # El valor no es un ID conocido (o ya era un nombre)
                    sin_match += 1
                    ids_sin_match.append(f"{col}:fila {idx + 1} → '{valor_str}'")

        # Reporte por archivo
        self.append_log(f"  ✔ Reemplazados:  {reemplazados}")
        self.append_log(f"  ○ Vacíos:        {vacias}")
        if sin_match:
            self.append_log(f"  ⚠ Sin coincidencia en VALIDACION: {sin_match}")
            for detalle in ids_sin_match:
                self.append_log(f"     {detalle}")

        # Guardar como _REVERTIDO (no pisa el _PROCESADO)
        ruta_salida = ruta_procesado.parent / (
            # Reemplaza el sufijo _PROCESADO por _REVERTIDO
            ruta_procesado.stem.replace("_PROCESADO", "")
            + "_REVERTIDO"
            + ruta_procesado.suffix
        )
        df.to_excel(ruta_salida, index=False)
        self.append_log(f"  ✔ Guardado en: {ruta_salida.name}")


if __name__ == "__main__":
    app = QApplication([])
    window = MainWindow()
    window.show()
    app.exec()
